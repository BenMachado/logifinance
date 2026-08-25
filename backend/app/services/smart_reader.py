"""Serviço de Leitura Inteligente de Arquivos.

Aceita planilhas (xlsx/xls/csv), PDFs, e imagens (jpg/png/webp) e devolve
uma lista estruturada de ``SmartReadItem`` pronta para o usuário revisar
e importar via POST /api/v1/costs.

Heurísticas:
- Planilha: tenta identificar cabeçalhos comuns (data/valor/descrição/categoria)
  e extrai cada linha como item.
- PDF: usa PyMuPDF para extrair texto; cada "linha" com valor monetário
  vira um item.
- Imagem: reusa o ``ocr_service`` já existente (Tesseract) e classifica o
  conteúdo como recibo ou texto livre.
"""
from __future__ import annotations

import csv
import io
import re
from dataclasses import dataclass
from datetime import date as date_type, datetime
from decimal import Decimal
from pathlib import Path
from typing import Iterable, List, Optional, Tuple

from app.models.cost_entry import CostCategory
from app.schemas.smart_read import SmartReadItem, SmartReadSummary
from app.services.ocr_service import (
    detect_amount,
    detect_category,
    detect_plate,
    parse_brazilian_amount,
)

# --- Classificadores & heurísticas ---------------------------------------

# Cabeçalhos comuns em planilhas financeiras (PT-BR).
_HEADER_KEYS = {
    "data": "date",
    "date": "date",
    "dt": "date",
    "vencimento": "date",
    "data_v": "date",
    "valor": "amount",
    "vlr": "amount",
    "amount": "amount",
    "preco": "amount",
    "preço": "amount",
    "total": "amount",
    "descricao": "description",
    "descrição": "description",
    "desc": "description",
    "historico": "description",
    "histórico": "description",
    "categoria": "category",
    "tipo": "category",
    "placa": "plate",
    "veiculo": "plate",
    "veículo": "plate",
}

_CATEGORY_KEYWORDS: list[tuple[CostCategory, tuple[str, ...]]] = [
    (CostCategory.FUEL, ("posto", "combust", "diesel", "gasolina", "etanol", "litro", "l.")),
    (CostCategory.TOLL, ("pedagio", "pedágio", "praca", "praça", "tag", "concessionaria", "concessionária", "ccr", "eco", "via")),
    (CostCategory.MAINTENANCE, ("oficina", "mecan", "pneu", "oleo", "óleo", "manutencao", "manutenção", "troca", "revisao", "revisão", "freio")),
    (CostCategory.FOOD, ("refeicao", "refeição", "almoco", "almoço", "janta", "jantar", "hotel", "diária", "diaria")),
    (CostCategory.INSURANCE, ("seguro", "apolice", "apólice")),
    (CostCategory.TAX, ("imposto", "taxa", "ipva", "licenciamento", "multa")),
    (CostCategory.MAINTENANCE, ("multa",)),  # multas viram manutenção/custo extra
]

_DATE_PATTERNS: list[re.Pattern[str]] = [
    re.compile(r"\b(\d{4})-(\d{2})-(\d{2})\b"),  # ISO
    re.compile(r"\b(\d{2})/(\d{2})/(\d{4})\b"),  # BR
    re.compile(r"\b(\d{2})-(\d{2})-(\d{4})\b"),  # BR com -
]


# --- Modelo interno ------------------------------------------------------


@dataclass
class ExtractedItem:
    line: int
    amount: Optional[Decimal]
    incurred_on: Optional[date_type]
    category: CostCategory
    description: Optional[str]
    plate: Optional[str]
    raw_text: Optional[str]
    confidence: float


# --- Helpers -------------------------------------------------------------


def _normalize_header(label: str) -> Optional[str]:
    if not label:
        return None
    key = (
        label.strip()
        .lower()
        .replace("ç", "c")
        .replace("ã", "a")
        .replace("á", "a")
        .replace("é", "e")
        .replace("í", "i")
        .replace("ó", "o")
        .replace("ú", "u")
    )
    key = re.sub(r"[^a-z0-9_]+", "", key)
    return _HEADER_KEYS.get(key)


def _classify_from_text(text: str) -> CostCategory:
    if not text:
        return CostCategory.OTHER
    t = text.lower()
    for cat, keywords in _CATEGORY_KEYWORDS:
        for kw in keywords:
            if kw in t:
                return cat
    # Fallback para o classificador OCR existente
    legacy = detect_category(text)
    try:
        return CostCategory(legacy)
    except ValueError:
        return CostCategory.OTHER


def _extract_date(text: str) -> Optional[date_type]:
    if not text:
        return None
    for pat in _DATE_PATTERNS:
        m = pat.search(text)
        if not m:
            continue
        try:
            if pat.pattern.startswith(r"\b(\d{4})"):
                y, mo, d = int(m.group(1)), int(m.group(2)), int(m.group(3))
            else:
                d, mo, y = int(m.group(1)), int(m.group(2)), int(m.group(3))
            return date_type(y, mo, d)
        except ValueError:
            continue
    return None


def _extract_amount(text: str) -> Optional[Decimal]:
    if not text:
        return None
    cleaned = str(text).strip()
    # Primeiro tenta direto
    direct = parse_brazilian_amount(cleaned)
    if direct is not None:
        return direct
    return detect_amount(cleaned)


def _extract_plate(text: str) -> Optional[str]:
    return detect_plate(text or "")


def _to_item_dict(item: ExtractedItem) -> SmartReadItem:
    return SmartReadItem(
        line=item.line,
        amount=item.amount,
        incurred_on=item.incurred_on,
        category=item.category,
        description=item.description,
        plate=item.plate,
        raw_text=item.raw_text,
        confidence=round(item.confidence, 2),
    )


def _build_summary(items: list[SmartReadItem]) -> SmartReadSummary:
    by_cat: dict[str, int] = {}
    total = Decimal("0")
    with_amount = 0
    with_date = 0
    for it in items:
        by_cat[it.category.value] = by_cat.get(it.category.value, 0) + 1
        if it.amount is not None:
            total += it.amount
            with_amount += 1
        if it.incurred_on is not None:
            with_date += 1
    return SmartReadSummary(
        total_items=len(items),
        items_with_amount=with_amount,
        items_with_date=with_date,
        total_amount=total if with_amount else None,
        by_category=by_cat,
    )


def _build_message_summary(text: str, file_type: str) -> Optional[str]:
    """Gera um resumo curto em PT-BR do texto extraído."""
    if not text:
        return None
    text = text.strip()
    if file_type == "image":
        n_lines = len([ln for ln in text.splitlines() if ln.strip()])
        return f"Texto OCR extraído com {n_lines} linhas. Revise os valores antes de importar."
    if file_type == "pdf":
        n_lines = len([ln for ln in text.splitlines() if ln.strip()])
        return f"PDF processado — {n_lines} linhas de texto. Itens com valor monetário foram extraídos."
    if file_type == "spreadsheet":
        return "Planilha lida — revise as categorias e datas sugeridas antes de importar."
    return None


# --- Detectores de tipo --------------------------------------------------


def detect_file_type(filename: str, content_type: Optional[str] = None) -> Tuple[str, str]:
    """Retorna (file_type, detected_format)."""
    suffix = (filename or "").lower().rsplit(".", 1)[-1] if "." in (filename or "") else ""
    if suffix in ("xlsx", "xls"):
        return "spreadsheet", suffix
    if suffix == "csv":
        return "spreadsheet", "csv"
    if suffix == "pdf":
        return "pdf", "pdf"
    if suffix in ("jpg", "jpeg", "png", "webp"):
        return "image", suffix
    if content_type:
        ct = content_type.lower()
        if ct.startswith("image/"):
            return "image", ct.split("/", 1)[1]
        if ct == "application/pdf":
            return "pdf", "pdf"
        if "spreadsheet" in ct or ct in ("text/csv", "application/vnd.ms-excel"):
            return "spreadsheet", "csv"
    return "unknown", suffix or "unknown"


# --- Parsers -------------------------------------------------------------


def parse_xlsx(path: Path) -> list[ExtractedItem]:
    try:
        from openpyxl import load_workbook  # type: ignore
    except ImportError as exc:  # pragma: no cover
        raise RuntimeError("openpyxl não instalado. Adicione ao requirements.txt.") from exc

    wb = load_workbook(filename=str(path), data_only=True, read_only=True)
    items: list[ExtractedItem] = []
    raw_text_parts: list[str] = []

    for sheet in wb.sheetnames:
        ws = wb[sheet]
        rows_iter = ws.iter_rows(values_only=True)
        try:
            header_row = next(rows_iter)
        except StopIteration:
            continue

        header_map: list[Optional[str]] = [_normalize_header(str(c) if c is not None else "") for c in header_row]

        for idx, row in enumerate(rows_iter, start=2):
            if row is None or all(c is None or (isinstance(c, str) and not c.strip()) for c in row):
                continue

            cells = [c if c is not None else "" for c in row]
            line_text = " | ".join(str(c) for c in cells if str(c).strip())
            if not line_text:
                continue
            raw_text_parts.append(line_text)

            amount: Optional[Decimal] = None
            date_val: Optional[date_type] = None
            description: Optional[str] = None
            plate: Optional[str] = None

            for col_idx, value in enumerate(cells):
                field = header_map[col_idx] if col_idx < len(header_map) else None
                if value is None or value == "":
                    continue
                if field == "amount":
                    amount = _extract_amount(str(value)) or amount
                elif field == "date":
                    if isinstance(value, datetime):
                        date_val = value.date()
                    elif isinstance(value, date_type):
                        date_val = value
                    else:
                        date_val = _extract_date(str(value)) or date_val
                elif field == "description":
                    description = str(value).strip()[:255]
                elif field == "plate":
                    plate = _extract_plate(str(value))
                else:
                    # Sem header mapeado — tenta extrair do conteúdo
                    if amount is None:
                        amount = _extract_amount(str(value))
                    if date_val is None:
                        if isinstance(value, datetime):
                            date_val = value.date()
                        elif isinstance(value, date_type):
                            date_val = value
                        else:
                            date_val = _extract_date(str(value))

            if amount is None and date_val is None and not description:
                # Linha provavelmente é cabeçalho/categoria sem dados úteis
                continue

            # Placa: se não veio de coluna, tenta na linha inteira
            if plate is None:
                plate = _extract_plate(line_text)
            # Descrição: fallback se vazia
            if not description:
                description = line_text[:255]

            category = _classify_from_text(" ".join(filter(None, [description, line_text])))

            items.append(
                ExtractedItem(
                    line=idx,
                    amount=amount,
                    incurred_on=date_val,
                    category=category,
                    description=description,
                    plate=plate,
                    raw_text=line_text[:1000],
                    confidence=0.8 if amount is not None and date_val is not None else 0.5,
                )
            )

    wb.close()
    # Anexa o texto cru para o resumo geral
    setattr(parse_xlsx, "_raw_text", "\n".join(raw_text_parts)[:4000])
    return items


def parse_csv(path: Path) -> list[ExtractedItem]:
    items: list[ExtractedItem] = []
    raw_text_parts: list[str] = []
    with path.open("r", encoding="utf-8-sig", errors="ignore", newline="") as fp:
        reader = csv.reader(fp)
        try:
            header_row = next(reader)
        except StopIteration:
            return items

        header_map: list[Optional[str]] = [_normalize_header(c) for c in header_row]

        for idx, row in enumerate(reader, start=2):
            if not row or all(not (c or "").strip() for c in row):
                continue
            line_text = " | ".join(c for c in row if (c or "").strip())
            raw_text_parts.append(line_text)

            amount: Optional[Decimal] = None
            date_val: Optional[date_type] = None
            description: Optional[str] = None
            plate: Optional[str] = None

            for col_idx, value in enumerate(row):
                field = header_map[col_idx] if col_idx < len(header_map) else None
                if not value:
                    continue
                if field == "amount":
                    amount = _extract_amount(value) or amount
                elif field == "date":
                    date_val = _extract_date(value) or date_val
                elif field == "description":
                    description = value.strip()[:255]
                elif field == "plate":
                    plate = _extract_plate(value)
                else:
                    if amount is None:
                        amount = _extract_amount(value)
                    if date_val is None:
                        date_val = _extract_date(value)

            if plate is None:
                plate = _extract_plate(line_text)
            if not description:
                description = line_text[:255]

            category = _classify_from_text(line_text)

            items.append(
                ExtractedItem(
                    line=idx,
                    amount=amount,
                    incurred_on=date_val,
                    category=category,
                    description=description,
                    plate=plate,
                    raw_text=line_text[:1000],
                    confidence=0.8 if amount is not None and date_val is not None else 0.5,
                )
            )

    setattr(parse_csv, "_raw_text", "\n".join(raw_text_parts)[:4000])
    return items


def parse_pdf(path: Path) -> list[ExtractedItem]:
    try:
        import fitz  # type: ignore  # PyMuPDF
    except ImportError as exc:  # pragma: no cover
        raise RuntimeError("pymupdf não instalado. Adicione ao requirements.txt.") from exc

    doc = fitz.open(str(path))
    items: list[ExtractedItem] = []
    raw_text_parts: list[str] = []

    for page_idx, page in enumerate(doc, start=1):
        text = page.get_text("text") or ""
        if text.strip():
            raw_text_parts.append(text)

        # 1) Tabelas — extrai blocos e cada linha vira um item se tiver valor.
        try:
            tables = page.find_tables()
            for t_idx, table in enumerate(tables):
                df = table.extract()
                for r_idx, row in enumerate(df):
                    cells = [str(c).strip() if c is not None else "" for c in row]
                    line_text = " | ".join(c for c in cells if c)
                    if not line_text:
                        continue
                    amount = detect_amount(line_text)
                    date_val = _extract_date(line_text)
                    if amount is None and date_val is None:
                        continue
                    plate = _extract_plate(line_text)
                    description = line_text[:255]
                    category = _classify_from_text(line_text)
                    items.append(
                        ExtractedItem(
                            line=page_idx * 1000 + t_idx * 100 + r_idx,
                            amount=amount,
                            incurred_on=date_val,
                            category=category,
                            description=description,
                            plate=plate,
                            raw_text=line_text[:1000],
                            confidence=0.7 if amount else 0.4,
                        )
                    )
        except Exception:
            pass

        # 2) Texto corrido — cada linha com R$/valor vira um item.
        for l_idx, line in enumerate(text.splitlines(), start=1):
            s = line.strip()
            if not s:
                continue
            amount = detect_amount(s)
            if amount is None:
                continue
            date_val = _extract_date(s)
            plate = _extract_plate(s)
            category = _classify_from_text(s)
            # Evita duplicar linhas que já estão em uma tabela — heurística simples
            already = any(it.raw_text == s[:1000] for it in items)
            if already:
                continue
            items.append(
                ExtractedItem(
                    line=page_idx * 10000 + l_idx,
                    amount=amount,
                    incurred_on=date_val,
                    category=category,
                    description=s[:255],
                    plate=plate,
                    raw_text=s[:1000],
                    confidence=0.6,
                )
            )

    doc.close()
    setattr(parse_pdf, "_raw_text", "\n".join(raw_text_parts)[:4000])
    return items


def parse_image(path: Path) -> tuple[list[ExtractedItem], str]:
    """Retorna (items, raw_text). Para imagens geralmente 1 item (recibo)."""
    from app.services.ocr_service import get_ocr_service

    text = get_ocr_service().extract(path)
    amount = detect_amount(text)
    plate = detect_plate(text)
    category_str = detect_category(text)
    try:
        category = CostCategory(category_str)
    except ValueError:
        category = CostCategory.OTHER

    items: list[ExtractedItem] = []
    if amount is not None or plate is not None:
        # Procura data no texto — senão usa a data de hoje
        incurred_on = _extract_date(text) or date_type.today()
        items.append(
            ExtractedItem(
                line=1,
                amount=amount,
                incurred_on=incurred_on,
                category=category,
                description=text.strip()[:255] or "Recibo OCR",
                plate=plate,
                raw_text=text[:1000],
                confidence=0.7 if amount else 0.4,
            )
        )
    return items, text


# --- Entry point ---------------------------------------------------------


def read_file(path: Path, file_type: str, detected_format: str) -> Tuple[List[SmartReadItem], SmartReadSummary, Optional[str], Optional[str]]:
    """Lê o arquivo e devolve (items, summary, message_summary, raw_text_preview)."""
    raw_text: Optional[str] = None

    if file_type == "spreadsheet":
        if detected_format == "csv":
            items = parse_csv(path)
            raw_text = getattr(parse_csv, "_raw_text", None)
        else:
            items = parse_xlsx(path)
            raw_text = getattr(parse_xlsx, "_raw_text", None)
    elif file_type == "pdf":
        items = parse_pdf(path)
        raw_text = getattr(parse_pdf, "_raw_text", None)
    elif file_type == "image":
        items, text = parse_image(path)
        raw_text = text
    else:
        items = []

    smart_items = [_to_item_dict(i) for i in items]
    summary = _build_summary(smart_items)
    message_summary = _build_message_summary(raw_text or "", file_type)
    return smart_items, summary, message_summary, raw_text


def parse_remote_url(url: str) -> Tuple[str, str, Path]:
    """Baixa uma URL e salva num arquivo temporário. Retorna (file_type, detected_format, path)."""
    import tempfile
    import urllib.request

    file_type, detected_format = detect_file_type(url)
    suffix = f".{detected_format}" if detected_format != "unknown" else ""
    fd, tmp_path = tempfile.mkstemp(suffix=suffix or ".bin")
    with urllib.request.urlopen(url, timeout=15) as resp:  # nosec - URL externa
        data = resp.read()
    Path(tmp_path).write_bytes(data)
    return file_type, detected_format, Path(tmp_path)
